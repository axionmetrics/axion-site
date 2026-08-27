#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Axion Metrics — Results Reconciler · FULL (τοπικό, 3 σήματα)
============================================================
Δημοσιεύτηκε (Euronext) × Κατέβηκε (φάκελος) × Μπήκε (RAW master) → radar.
ΔΕΝ γράφει τίποτα. Εξαιρεί frozen (UPDATED=NO) + μελλοντικές ημ/νίες, ΚΑΙ
δείχνει σε κάθε radar το «Εκτός κάλυψης (frozen)» roster.
"""
import json, re, os, sys, argparse, datetime
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import results_reconciler as rc

RAW_STRIDE = 37; RREP_OFFSETS = (9, 18, 26); BASE_YEAR = 2020
def raw_col(year): return 16 + (year - BASE_YEAR)
CODE_RE = re.compile(r'^\s*(\d+)'); YEAR_RE = re.compile(r'(20\d{2})')

# ---------------------------------------------------------------- folder
def folder_index_from_entries(entries, want_ext=(".pdf",)):
    idx = {}
    for e in entries:
        if e.get("type") != "file": continue
        name = e.get("name", "")
        if want_ext and not name.lower().endswith(tuple(want_ext)): continue
        mc = CODE_RE.match(name.split("/", 1)[0])
        if not mc: continue
        for y in YEAR_RE.findall(name.rsplit("/", 1)[-1]):
            idx.setdefault(mc.group(1), set()).add(y)
    return idx

def load_downloaded(args, basis):
    if args.folder_listing:
        entries = json.load(open(args.folder_listing, encoding="utf-8"))
        if isinstance(entries, dict): entries = entries.get("entries", [])
        return folder_index_from_entries(entries)
    if args.downloaded_json:
        d = json.load(open(args.downloaded_json, encoding="utf-8")).get(basis, {})
        return {str(k): set(v) for k, v in d.items()}
    return {}

def downloaded_year(basis, period): return period[:4] if basis == "interim" else period

# ---------------------------------------------------------------- master
def _num(v):
    try: return None if v is None else float(v)
    except Exception: return None

def open_master(path):
    import openpyxl
    return openpyxl.load_workbook(path, data_only=True)

def master_reported_by_tk(wb, basis, period):
    sh = wb["RAW H1" if basis == "interim" else "RAW"]
    yr = int(period[:4]) if basis == "interim" else int(period)
    wcol = raw_col(yr); out = set()
    for k in range(160):
        rr = 3 + RAW_STRIDE * k
        if sh.cell(rr, 3).value is None: continue
        tk = sh.cell(rr, 4).value; o0 = 2 + RAW_STRIDE * k
        if any(_num(sh.cell(o0 + off, wcol).value) not in (None, 0) for off in RREP_OFFSETS) and tk:
            out.add(str(tk).strip())
    return out

def read_index(wb):
    """-> links{code:url}, excluded_tks{set}, frozen_roster[{code,name,note}]."""
    links, ex_tk, roster = {}, set(), []
    if "INDEX ΕΠΙΧΕΙΡΗΣΕΩΝ" not in wb.sheetnames:
        return links, ex_tk, roster
    idx = wb["INDEX ΕΠΙΧΕΙΡΗΣΕΩΝ"]; lcol = jcol = ncol = None
    for c in range(1, idx.max_column + 1):
        h = str(idx.cell(3, c).value or "").strip().upper()
        if h.startswith("ΣΕΛΙΔΑ ΟΙΚΟΝΟΜΙΚ"): lcol = c
        elif h == "UPDATED" or h == "CALCULATED": jcol = c
        elif h == "ΣΗΜΕΙΩΣΗ ΟΝΤΟΤΗΤΑΣ": ncol = c
    for r in range(4, idx.max_row + 1):
        a = idx.cell(r, 1).value
        if not a: continue
        mc = CODE_RE.match(str(a))
        if not mc: continue
        code = mc.group(1)
        if lcol and idx.cell(r, lcol).value: links[code] = str(idx.cell(r, lcol).value).strip()
        if jcol and str(idx.cell(r, jcol).value or "").strip().upper() == "NO":
            tk = idx.cell(r, 3).value
            if tk: ex_tk.add(str(tk).strip())
            note = str(idx.cell(r, ncol).value or "").split("||")[0].strip() if ncol else ""
            roster.append({"code": code, "name": str(a).strip(), "note": note})
    roster.sort(key=lambda x: int(x["code"]))
    return links, ex_tk, roster

# ---------------------------------------------------------------- join
def full_rows(published, downloaded, reported_tks, links, basis, period):
    dyear = downloaded_year(basis, period); rows = []
    for r in published:
        code = str(r["code"])
        rows.append({**r, "downloaded": dyear in downloaded.get(code, set()),
                     "entered": r["tk"] in reported_tks, "link": links.get(code, "")})
    rows.sort(key=lambda x: (x["entered"], rc.date_key(x["date"])))
    return rows

def bucketize(rows):
    return ([r for r in rows if not r["entered"] and not r["downloaded"]],
            [r for r in rows if not r["entered"] and r["downloaded"]],
            [r for r in rows if r["entered"]])

# ---------------------------------------------------------------- render
def _short(n): return n.split(" ", 1)[1] if " " in n else n
def _sig(ok): return ('<span class="sig y">✓</span>' if ok else '<span class="sig n">✗</span>')
def _lab(basis, period): return (f"6μηνο {period[:4]}" if basis == "interim" else f"έτος {period}")

def render_html(results, asof, roster):
    css = """*{box-sizing:border-box}body{margin:0;background:#f5f7fa;color:#1b2740;font-family:Inter,system-ui,Arial,sans-serif;font-size:14px;line-height:1.45}
    .wrap{max-width:1080px;margin:0 auto;padding:26px 20px 60px}h1{font-size:23px;font-weight:800;margin:0 0 3px}
    .sub{color:#8a95a6;font-size:13px;margin-bottom:20px}
    h2{font-size:13px;letter-spacing:.05em;text-transform:uppercase;color:#8a95a6;margin:26px 0 9px}
    .chips{display:flex;gap:12px;flex-wrap:wrap;margin:0 0 8px}
    .chip{background:#fff;border:1px solid #e7ebf2;border-radius:11px;padding:10px 15px;min-width:150px;box-shadow:0 2px 8px rgba(27,39,64,.04)}
    .chip .n{font-size:23px;font-weight:800;line-height:1}.chip .l{font-size:11px;color:#8a95a6;margin-top:3px}.chip.hot .n{color:#c0392b}.chip.warn .n{color:#b8860b}
    .tblwrap{overflow-x:auto;background:#fff;border:1px solid #e7ebf2;border-radius:12px;box-shadow:0 3px 14px rgba(27,39,64,.05)}
    table{border-collapse:collapse;width:100%;min-width:760px}th,td{text-align:left;padding:10px 13px;border-bottom:1px solid #eef1f6;white-space:nowrap}
    th{font-size:10.5px;letter-spacing:.05em;text-transform:uppercase;color:#8a95a6;background:#fafbfc}tr:last-child td{border-bottom:0}
    .co{font-weight:700}.tk{color:#8a95a6;font-size:12px}
    .sig{display:inline-flex;align-items:center;justify-content:center;width:22px;height:22px;border-radius:50%;font-size:12px;font-weight:800}
    .sig.y{background:rgba(15,138,77,.13);color:#0f8a4d}.sig.n{background:rgba(192,57,43,.11);color:#c0392b}
    .act{font-weight:600}.act.dl{color:#c0392b}.act.raw{color:#b8860b}.act.ok{color:#0f8a4d}
    a.dl{color:#2f80c2;text-decoration:none;font-weight:600;font-size:12.5px}a.dl:hover{text-decoration:underline}
    .done td{opacity:.6}.fz td{opacity:.7}.fzbadge{display:inline-block;background:#fdf4e3;color:#b57d00;border:1px solid #f0d8a6;border-radius:20px;font-size:10.5px;font-weight:700;padding:1px 8px;margin-left:7px}
    .foot{color:#8a95a6;font-size:12px;margin-top:18px;border-top:1px solid #e7ebf2;padding-top:12px}
    code{background:#eef1f6;border-radius:4px;padding:1px 5px;font-size:12px}"""
    H = [f"<title>AXION · Results Radar</title><style>{css}</style>", '<div class="wrap">',
         "<h1>AXION · Results Radar</h1>",
         f'<div class="sub">FULL 3-σημάτων · {asof} · Euronext × φάκελος × master RAW · φίλτρο frozen</div>']
    for basis, period, rows in results:
        todo_dl, todo_raw, done = bucketize(rows)
        H.append(f"<h2>{_lab(basis, period)}</h2>")
        H.append('<div class="chips">'
                 f'<div class="chip hot"><div class="n">{len(todo_dl)}</div><div class="l">κατέβασε → RAW</div></div>'
                 f'<div class="chip warn"><div class="n">{len(todo_raw)}</div><div class="l">κατέβηκε — λείπει RAW</div></div>'
                 f'<div class="chip"><div class="n">{len(done)}</div><div class="l">ολοκληρωμένες (live)</div></div></div>')
        def tbl(title, rws, showlink=True):
            if not rws: return
            H.append(f'<h2 style="margin-top:16px">{title}</h2><div class="tblwrap"><table>')
            H.append("<tr><th>Εταιρεία</th><th>Ticker</th><th>Δημοσίευση</th><th>Δημοσ.</th><th>Κατέβ.</th><th>RAW</th><th>Ενέργεια</th>"
                     + ("<th>Καταστάσεις</th>" if showlink else "") + "</tr>")
            for r in rws:
                act = ('<span class="act ok">OK — live</span>' if r["entered"]
                       else '<span class="act raw">πέρασε στο RAW</span>' if r["downloaded"]
                       else '<span class="act dl">κατέβασε → RAW</span>')
                link = (f'<a class="dl" href="{r["link"]}">σελίδα ↗</a>' if r["link"] else "—") if showlink else ""
                cls = ' class="done"' if r["entered"] else ""
                H.append(f'<tr{cls}><td class="co">{r["code"]} · {_short(r["name"])}</td><td class="tk">{r["tk"]}</td>'
                         f'<td>{r["date"]}</td><td>{_sig(True)}</td><td>{_sig(r["downloaded"])}</td><td>{_sig(r["entered"])}</td>'
                         f'<td class="act">{act}</td>' + (f"<td>{link}</td>" if showlink else "") + "</tr>")
            H.append("</table></div>")
        tbl("⟵ Προς ενέργεια — δεν κατέβηκαν", todo_dl)
        tbl("Κατέβηκαν — λείπει μόνο η καταχώρηση RAW", todo_raw)
        tbl("Ολοκληρωμένες (live)", done, showlink=False)
    # frozen roster — ΠΑΝΤΑ
    H.append(f'<h2>⏸️ Εκτός κάλυψης (frozen) — δεν παρακολουθούνται · {len(roster)}</h2>')
    if roster:
        H.append('<div class="tblwrap"><table><tr><th>Εταιρεία</th><th>Αιτία</th></tr>')
        for r in roster:
            H.append(f'<tr class="fz"><td class="co">{r["code"]} · {_short(r["name"])}'
                     f'<span class="fzbadge">Εκτός κάλυψης</span></td><td>{r["note"]}</td></tr>')
        H.append("</table></div>")
    else:
        H.append('<div class="sub">Καμία.</div>')
    H.append('<div class="foot">Σήματα: <b>Δημοσ.</b>=Euronext (≤ σήμερα) · <b>Κατέβ.</b>=PDF περιόδου στον φάκελο · '
             '<b>RAW</b>=<code>reported</code> στο master. Οι frozen (UPDATED=NO) εξαιρούνται από τη λίστα εργασίας. '
             'Δεν γράφει τίποτα.</div></div>')
    return "\n".join(H)

def render_md(results, asof, roster):
    L = [f"# 🎯 Results Radar (FULL) — {asof}", ""]
    for basis, period, rows in results:
        todo_dl, todo_raw, done = bucketize(rows)
        L += [f"## {_lab(basis, period)}",
              f"κατέβασε→RAW: **{len(todo_dl)}** · κατέβηκε-λείπει-RAW: **{len(todo_raw)}** · live: **{len(done)}**", ""]
        def tbl(title, rws):
            if not rws: return
            L.append(f"**{title}**")
            L.extend(["| Εταιρεία | Ticker | Δημοσ. | Κατέβ. | RAW | Ενέργεια | Καταστάσεις |","|---|---|---|---|---|---|---|"])
            for r in rws:
                act = "OK — live" if r["entered"] else ("→ RAW" if r["downloaded"] else "κατέβασε → RAW")
                link = f"[σελίδα]({r['link']})" if r["link"] else "—"
                L.append(f"| {r['code']} · {_short(r['name'])} | {r['tk']} | {r['date']} | ✓ | "
                         f"{'✓' if r['downloaded'] else '✗'} | {'✓' if r['entered'] else '✗'} | {act} | {link} |")
            L.append("")
        tbl("⟵ Προς ενέργεια — δεν κατέβηκαν", todo_dl)
        tbl("Κατέβηκαν — λείπει μόνο RAW", todo_raw)
        tbl("Ολοκληρωμένες (live)", done)
    L += [f"## ⏸️ Εκτός κάλυψης (frozen) · {len(roster)}", ""]
    L += ([f"- {r['code']} · {_short(r['name'])} — {r['note']}" for r in roster] if roster else ["_Καμία._"])
    return "\n".join(L)

# ---------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--cid-map", default="euronext_cid_map.json")
    ap.add_argument("--master", required=True)
    ap.add_argument("--events-json", required=True)
    ap.add_argument("--downloaded-json"); ap.add_argument("--folder-listing")
    ap.add_argument("--basis", default="interim", choices=["interim", "annual"])
    ap.add_argument("--both", action="store_true")
    ap.add_argument("--period"); ap.add_argument("--asof")
    ap.add_argument("--out-html"); ap.add_argument("--out-md")
    args = ap.parse_args()
    asof = datetime.date.fromisoformat(args.asof) if args.asof else datetime.datetime.utcnow().date()
    cid2row, _ = rc.load_cid_map(args.cid_map)
    events = json.load(open(args.events_json, encoding="utf-8"))
    wb = open_master(args.master)
    links, excluded_tks, roster = read_index(wb)
    bases = ["interim", "annual"] if args.both else [args.basis]
    results = []
    for b in bases:
        period = args.period if (args.period and not args.both) else \
                 (f"{asof.year}H1" if b == "interim" else str(asof.year - 1))
        pub = rc.reconcile(events, cid2row, set(), b, period, excluded_tks=excluded_tks, asof=asof)["published"]
        rows = full_rows(pub, load_downloaded(args, b), master_reported_by_tk(wb, b, period), links, b, period)
        results.append((b, period, rows))
    html = render_html(results, asof, roster); md = render_md(results, asof, roster)
    if args.out_html: open(args.out_html, "w", encoding="utf-8").write(html)
    if args.out_md:   open(args.out_md, "w", encoding="utf-8").write(md + "\n")
    print(md)

if __name__ == "__main__":
    main()
