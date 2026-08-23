#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
apply_events.py — Axion Metrics
================================
Περνά τα γεγονότα από την ουρά του monitor (.github/monitor/events_queue.csv)
στο master «ΑΡΙΘΜΟΔΕΙΚΤΕΣ.xlsx», φύλλο «ΕΤΑΙΡΙΚΑ ΓΕΓΟΝΟΤΑ».

ΤΡΕΧΕΙ ΤΟΠΙΚΑ (όχι στο GitHub Action) — γράφει το τοπικό master.
Ο monitor τρέχει στο cloud και ΔΕΝ αγγίζει το master· εδώ γίνεται η εγγραφή.

Τι κάνει, ανά γεγονός της ουράς:
  • Βρίσκει το εσωτερικό ticker από το euronext_symbol μέσω INDEX ΕΠΙΧΕΙΡΗΣΕΩΝ
    (ταιριάζει σε ENGLISH ticker / GREEK ticker / ISIN).
  • Χτίζει γραμμή A..N με τη σωστή σειρά στηλών του master.
  • Συμπληρώνει M «Ποσό €/μτχ» (από amount_eur) και N «Χρήση» (από fiscal_year).
  • Dedup: δεν ξαναγράφει γεγονός που υπάρχει ήδη (ticker+ημ/νία+τύπος+ποσό).
  • Ό,τι δεν λύνεται (άγνωστο symbol ή κενή/ασαφής χρήση) -> report «needs_detail»
    ΔΕΝ γράφεται· το τακτοποιείς με το χέρι.

⚠ Μετά το apply: ΑΝΟΙΞΕ+ΣΩΣΕ το master στο Excel πριν τρέξει η γέφυρα
  (openpyxl αφαιρεί cached τιμές· το SUMIFS του ΟΝΟΜΑΣΤΙΚΟΥ ξαναϋπολογίζεται στο Excel).

Χρήση:
  python apply_events.py --master "ΑΡΙΘΜΟΔΕΙΚΤΕΣ.xlsx" --queue events_queue.csv [--dry-run]
"""
import argparse, csv, copy, datetime, os, sys, re

try:
    import openpyxl
except ImportError:
    print("Λείπει το openpyxl (pip install openpyxl).", file=sys.stderr); raise

EVSHEET = "ΕΤΑΙΡΙΚΑ ΓΕΓΟΝΟΤΑ"
IDXSHEET = "INDEX ΕΠΙΧΕΙΡΗΣΕΩΝ"
# Στήλες master (1-based): A..N
COL = dict(aa=1, company=2, ticker=3, date=4, year=5, category=6, family=7,
           type=8, desc=9, status=10, srctitle=11, source=12, amount=13, fiscal=14)


def norm(s):
    return (str(s).strip() if s is not None else "")


def build_resolver(wb):
    """euronext symbol/ISIN -> εσωτερικό ticker (INDEX col C 'GREEK TICKER')."""
    idx = wb[IDXSHEET]
    res = {}
    name = {}
    # headers στη row 3· δεδομένα από row 4
    for r in range(4, idx.max_row + 1):
        gr = norm(idx.cell(r, 3).value)   # C GREEK TICKER (εσωτερικό)
        en = norm(idx.cell(r, 5).value)   # E ENGLISH TICKER (Euronext symbol)
        isin = norm(idx.cell(r, 4).value)  # D ISIN
        dbn = norm(idx.cell(r, 1).value)  # A ΟΝΟΜΑ ΣΤΗ ΒΑΣΗ ΜΟΥ
        if not gr:
            continue
        for keyv in (gr, en, isin):
            if keyv:
                res.setdefault(keyv.upper(), gr)
        name[gr] = dbn or gr
    return res, name


def existing_keys(ws):
    """Κλειδιά υπαρχόντων γεγονότων για dedup: ticker|iso-date|type|amount."""
    keys = set()
    for r in range(2, ws.max_row + 1):
        tk = norm(ws.cell(r, COL["ticker"]).value)
        if not tk:
            continue
        d = ws.cell(r, COL["date"]).value
        diso = d.strftime("%Y-%m-%d") if hasattr(d, "strftime") else norm(d)
        typ = norm(ws.cell(r, COL["type"]).value)
        amt = ws.cell(r, COL["amount"]).value
        keys.add("%s|%s|%s|%s" % (tk.upper(), diso, typ, amt))
    return keys


def to_date(iso):
    try:
        y, m, d = iso.split("-")
        return datetime.datetime(int(y), int(m), int(d))
    except Exception:
        return iso


def to_num(s):
    s = norm(s)
    if not s:
        return None
    s = s.strip(".,")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


def to_year(s):
    m = re.search(r"(19|20)\d{2}", norm(s))
    return int(m.group(0)) if m else None


def style_from(ws, src_row):
    """Επιστρέφει (font,fill,border,align,number_format) ανά στήλη από γραμμή-πρότυπο."""
    st = {}
    for c in range(1, COL["fiscal"] + 1):
        cell = ws.cell(src_row, c)
        st[c] = (copy.copy(cell.font), copy.copy(cell.fill),
                 copy.copy(cell.border), copy.copy(cell.alignment),
                 cell.number_format)
    return st


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--master", required=True)
    ap.add_argument("--queue", required=True)
    ap.add_argument("--out", default=None, help="default: overwrite master path")
    ap.add_argument("--dry-run", action="store_true")
    a = ap.parse_args()

    wb = openpyxl.load_workbook(a.master, data_only=False)
    ws = wb[EVSHEET]
    resolver, dbname = build_resolver(wb)
    seen = existing_keys(ws)

    # γραμμή-πρότυπο στυλ = τελευταία γραμμή με δεδομένα
    last = ws.max_row
    while last > 1 and not norm(ws.cell(last, COL["ticker"]).value):
        last -= 1
    st = style_from(ws, last)
    # επόμενο Α/Α
    aa_vals = [ws.cell(r, COL["aa"]).value for r in range(2, last + 1)
               if isinstance(ws.cell(r, COL["aa"]).value, (int, float))]
    next_aa = int(max(aa_vals)) + 1 if aa_vals else 1

    added, skipped, needs = 0, [], []
    with open(a.queue, encoding="utf-8", newline="") as f:
        for q in csv.DictReader(f):
            sym = norm(q.get("euronext_symbol"))
            fam = norm(q.get("family"))
            tk = resolver.get(sym.upper())
            diso = norm(q.get("date"))
            typ = norm(q.get("type"))
            amt = to_num(q.get("amount_eur"))
            key = "%s|%s|%s|%s" % ((tk or sym).upper(), diso, typ, amt)
            if not tk:
                needs.append((sym, "άγνωστο symbol", q.get("description")))
                continue
            if key in seen:
                skipped.append(key)
                continue
            fiscal = to_year(q.get("fiscal_year")) if fam in ("div", "capital") else None
            if fam in ("div", "capital") and fiscal is None:
                # Γράφουμε ΚΑΝΟΝΙΚΑ τη διανομή (να φανεί στο γράφημα) αλλά με ΚΕΝΗ Χρήση,
                # και flag ώστε να ορίσεις εσύ τη χρήση (π.χ. έκτακτο χωρίς fiscal στο Euronext).
                needs.append((tk, "κενή Χρήση — όρισέ τη με το χέρι", q.get("description")))
            seen.add(key)
            if a.dry_run:
                added += 1
                continue
            # append
            r = ws.max_row + 1
            vals = {
                COL["aa"]: next_aa,
                COL["company"]: dbname.get(tk, tk),
                COL["ticker"]: tk,
                COL["date"]: to_date(diso),
                COL["year"]: (to_date(diso).year if hasattr(to_date(diso), "year") else None),
                COL["category"]: norm(q.get("category")),
                COL["family"]: fam,
                COL["type"]: typ,
                COL["desc"]: norm(q.get("description")),
                COL["status"]: "OK",
                COL["srctitle"]: norm(q.get("source_title")),
                COL["source"]: None,
                COL["amount"]: amt,
                COL["fiscal"]: fiscal,
            }
            for c in range(1, COL["fiscal"] + 1):
                cell = ws.cell(r, c, vals.get(c))
                fnt, fil, brd, aln, nf = st[c]
                cell.font, cell.fill, cell.border, cell.alignment = fnt, fil, brd, aln
                cell.number_format = nf
            next_aa += 1
            added += 1

    # ενημέρωσε autofilter ώστε να καλύπτει όλα τα δεδομένα
    lastdata = ws.max_row
    if ws.auto_filter.ref:
        mm = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", ws.auto_filter.ref)
        if mm:
            ws.auto_filter.ref = "%s%s:N%d" % (mm.group(1), mm.group(2), lastdata)

    print("apply_events: added=%d  skipped(dup)=%d  needs_detail=%d"
          % (added, len(skipped), len(needs)))
    for n in needs:
        print("  ! NEEDS:", n)

    if not a.dry_run and added:
        out = a.out or a.master
        wb.save(out)
        print("saved:", out)
        print("⚠ Άνοιξε+Σώσε το master στο Excel πριν τρέξει η γέφυρα.")


if __name__ == "__main__":
    main()
